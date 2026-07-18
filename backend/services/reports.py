import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession

from services.dashboard import DashboardService

_HEADER_FILL = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="111111")


def _write_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 50)


async def generate_sales_inventory_report(db: AsyncSession) -> bytes:
    """Builds a multi-sheet Excel workbook entirely from services/dashboard.py's existing
    aggregation methods — no parallel query logic, this only formats what DashboardService
    already computes for the dashboard UI into a downloadable report."""
    stats = await DashboardService.get_dashboard_stats(db)
    trend = await DashboardService.get_revenue_trend(db, days=30)
    top_products = await DashboardService.get_top_products(db, limit=20)
    low_stock = await DashboardService.get_low_stock_items(db, limit=100)

    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet — every sheet below is named explicitly

    _write_sheet(wb, "Summary", ["Metric", "Value"], [
        ["Total Sales (delivered orders)", stats["total_sales"]],
        ["Total Orders", stats["total_orders"]],
        ["Total Users", stats["total_users"]],
        ["Pending Orders", stats["pending_orders"]],
        ["Low Stock Items", stats["low_stock_items"]],
        ["Revenue Today", stats["revenue_today"]],
        ["Orders Today", stats["orders_today"]],
    ])

    _write_sheet(wb, "Revenue Trend (30d)", ["Date", "Revenue"], [
        [label, value] for label, value in zip(trend["labels"], trend["data"])
    ])

    _write_sheet(wb, "Top Products", ["Product", "Units Sold", "Revenue"], [
        [p["name"], p["total_sold"], p["revenue"]] for p in top_products
    ])

    _write_sheet(wb, "Low Stock", ["Product", "Category", "Total Stock"], [
        [p["name"], p["category"], p["total_stock"]] for p in low_stock
    ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
